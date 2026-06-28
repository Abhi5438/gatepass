import os
import datetime
import hashlib
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, send_file, session
from openpyxl import Workbook, load_workbook

app = Flask(__name__, static_folder='static', static_url_path='')

# Secret key for Flask session signing
app.secret_key = 'gatepass_secure_session_key_2026_securpass'
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=7)

# Define constants
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gatepass_data.xlsx')
UPLOAD_FOLDER = os.path.join(app.static_folder, 'uploads')

# Ensure uploads folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Helper function to hash password using SHA-256
def hash_password(password):
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def init_excel():
    wb = None
    file_exists = os.path.exists(EXCEL_FILE)
    
    if not file_exists:
        wb = Workbook()
        ws_v = wb.active
        ws_v.title = "Vehicles"
    else:
        try:
            wb = load_workbook(EXCEL_FILE)
        except Exception as e:
            print(f"Error loading workbook, recreating: {e}")
            wb = Workbook()
            ws_v = wb.active
            ws_v.title = "Vehicles"
            
    # Setup Vehicles worksheet if not exists
    if "Vehicles" not in wb.sheetnames:
        ws_v = wb.create_sheet("Vehicles")
    else:
        ws_v = wb["Vehicles"]
        
    # Check if Vehicles headers are initialized
    if ws_v.max_row == 1 and ws_v.cell(row=1, column=1).value is None:
        headers = [
            "Vehicle No", "Driver Name", "Punch Number", "Gatepass Expiry Date", 
            "Insurance Expiry Date", "PUC Expiry Date", 
            "Fitness Expiry Date", "Tax Expiry Date", "Permit Expiry Date",
            "Inspection Stage", "Safety Stage", "Department Stage",
            "Remark", "DL Photo", "Punch Card", "Created At"
        ]
        for col_num, h in enumerate(headers, 1):
            ws_v.cell(row=1, column=col_num).value = h
        
        # Style headers (bold font, purple background fill)
        from openpyxl.styles import Font, PatternFill
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid") # Violet 800
        
        for col_num in range(1, len(headers) + 1):
            cell = ws_v.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            
        # Adjust column widths
        widths = [20, 25, 20, 15, 15, 15, 15, 15, 15, 20, 20, 20, 30, 30, 30, 20]
        for i, width in enumerate(widths):
            col_letter = ws_v.cell(row=1, column=i+1).column_letter
            ws_v.column_dimensions[col_letter].width = width

    # Setup Users worksheet if not exists
    if "Users" not in wb.sheetnames:
        ws_u = wb.create_sheet("Users")
        user_headers = ["Username", "Password Hash", "Role", "Status", "Created At"]
        for col_num, h in enumerate(user_headers, 1):
            ws_u.cell(row=1, column=col_num).value = h
        
        # Style User headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue 500
        for col_num in range(1, len(user_headers) + 1):
            cell = ws_u.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            
        # Seed default Admin user
        ws_u.append([
            "admin",
            hash_password("admin"),
            "admin",
            "approved",
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ])
        print("Users worksheet initialized and default admin seeded.")

    # Setup Drivers worksheet if not exists (including DL Number and Base Document)
    if "Drivers" not in wb.sheetnames:
        ws_d = wb.create_sheet("Drivers")
        drivers_headers = [
            "Punch Number", "Driver Name", "DL Number", "PAN Number", "Account Number",
            "DL Photo", "Punch Card Photo", "PAN Photo", "Passbook Photo", "Base Document",
            "Status", "Created At"
        ]
        for col_num, h in enumerate(drivers_headers, 1):
            ws_d.cell(row=1, column=col_num).value = h
        
        # Style Drivers headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald 500
        for col_num in range(1, len(drivers_headers) + 1):
            cell = ws_d.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            
        # Adjust column widths
        widths = [20, 25, 20, 20, 20, 30, 30, 30, 30, 30, 15, 20]
        for i, width in enumerate(widths):
            col_letter = ws_d.cell(row=1, column=i+1).column_letter
            ws_d.column_dimensions[col_letter].width = width
        print("Drivers worksheet initialized.")
    else:
        # Check if worksheet has DL Number column; if old style sheet, recreate to avoid structure mismatch
        ws_d = wb["Drivers"]
        if ws_d.max_column < 12:
            wb.remove(ws_d)
            ws_d = wb.create_sheet("Drivers")
            drivers_headers = [
                "Punch Number", "Driver Name", "DL Number", "PAN Number", "Account Number",
                "DL Photo", "Punch Card Photo", "PAN Photo", "Passbook Photo", "Base Document",
                "Status", "Created At"
            ]
            for col_num, h in enumerate(drivers_headers, 1):
                ws_d.cell(row=1, column=col_num).value = h
            
            # Style Drivers headers
            from openpyxl.styles import Font, PatternFill
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald 500
            for col_num in range(1, len(drivers_headers) + 1):
                cell = ws_d.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                
            # Adjust column widths
            widths = [20, 25, 20, 20, 20, 30, 30, 30, 30, 30, 15, 20]
            for i, width in enumerate(widths):
                col_letter = ws_d.cell(row=1, column=i+1).column_letter
                ws_d.column_dimensions[col_letter].width = width
            print("Drivers worksheet updated to DL Number & Base Doc layout.")

    wb.save(EXCEL_FILE)
    wb.close()
    print("Excel workbook check completed.")

# Helper queries for Vehicles
def get_vehicles():
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Vehicles"]
        vehicles = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: # Skip empty rows
                continue
            vehicles.append({
                "vehicleNo": str(row[0]).strip(),
                "driverName": str(row[1]).strip() if row[1] else "",
                "punchNumber": str(row[2]).strip() if row[2] else "",
                "gatepassExpiry": str(row[3]).strip() if row[3] else "",
                "insuranceExpiry": str(row[4]).strip() if row[4] else "",
                "pucExpiry": str(row[5]).strip() if row[5] else "",
                "fitnessExpiry": str(row[6]).strip() if row[6] else "",
                "taxExpiry": str(row[7]).strip() if row[7] else "",
                "permitExpiry": str(row[8]).strip() if row[8] else "",
                "stageInspection": str(row[9]).strip() if row[9] else "Pending",
                "stageSafety": str(row[10]).strip() if row[10] else "Pending",
                "stageDept": str(row[11]).strip() if row[11] else "Pending",
                "remark": str(row[12]).strip() if row[12] else "",
                "dlPhoto": str(row[13]).strip() if row[13] else "",
                "punchCard": str(row[14]).strip() if row[14] else "",
                "createdAt": str(row[15]).strip() if row[15] else ""
            })
        wb.close()
        return vehicles
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return []

def add_vehicle(v):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Vehicles"]
    ws.append([
        v["vehicleNo"],
        v["driverName"],
        v["punchNumber"],
        v["gatepassExpiry"],
        v["insuranceExpiry"],
        v["pucExpiry"],
        v["fitnessExpiry"],
        v["taxExpiry"],
        v["permitExpiry"],
        v.get("stageInspection", "Pending"),
        v.get("stageSafety", "Pending"),
        v.get("stageDept", "Pending"),
        v["remark"],
        v["dlPhoto"],
        v["punchCard"],
        v["createdAt"]
    ])
    wb.save(EXCEL_FILE)
    wb.close()

def delete_vehicle(vehicle_no):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Vehicles"]
    row_idx = -1
    dl_photo = ""
    punch_card = ""
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().upper() == vehicle_no.strip().upper():
            row_idx = r_idx
            dl_photo = ws.cell(row=r_idx, column=14).value or ""  # Column N (14)
            punch_card = ws.cell(row=r_idx, column=15).value or ""  # Column O (15)
            break
            
    if row_idx != -1:
        ws.delete_rows(row_idx)
        wb.save(EXCEL_FILE)
        wb.close()
        return True, dl_photo, punch_card
    wb.close()
    return False, "", ""

def toggle_vehicle_stage(vehicle_no, stage_name):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Vehicles"]
    updated = False
    new_val = ""
    col_idx = 10 # Default for Inspection (Column J)
    if stage_name == "Safety":
        col_idx = 11 # Column K
    elif stage_name == "Department":
        col_idx = 12 # Column L
        
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().upper() == vehicle_no.strip().upper():
            current_val = ws.cell(row=r_idx, column=col_idx).value or "Pending"
            new_val = "Pending" if current_val.strip() == "Done" else "Done"
            ws.cell(row=r_idx, column=col_idx).value = new_val
            updated = True
            break
    if updated:
        wb.save(EXCEL_FILE)
    wb.close()
    return updated, new_val

def update_vehicle_in_excel(vehicle_no, u):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Vehicles"]
    updated = False
    old_dl = ""
    old_pc = ""
    
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().upper() == vehicle_no.strip().upper():
            # Get old photo filenames to unlink from disk
            old_dl = ws.cell(row=r_idx, column=14).value or ""
            old_pc = ws.cell(row=r_idx, column=15).value or ""
            
            # Update basic details
            ws.cell(row=r_idx, column=2).value = u["driverName"]
            ws.cell(row=r_idx, column=3).value = u["punchNumber"]
            ws.cell(row=r_idx, column=4).value = u["gatepassExpiry"]
            ws.cell(row=r_idx, column=5).value = u["insuranceExpiry"]
            ws.cell(row=r_idx, column=6).value = u["pucExpiry"]
            ws.cell(row=r_idx, column=7).value = u["fitnessExpiry"]
            ws.cell(row=r_idx, column=8).value = u["taxExpiry"]
            ws.cell(row=r_idx, column=9).value = u["permitExpiry"]
            
            # Update stages
            ws.cell(row=r_idx, column=10).value = u["stageInspection"]
            ws.cell(row=r_idx, column=11).value = u["stageSafety"]
            ws.cell(row=r_idx, column=12).value = u["stageDept"]
            
            # Update remark
            ws.cell(row=r_idx, column=13).value = u["remark"]
            
            # If new files are uploaded, overwrite values
            if u["dlPhoto"]:
                ws.cell(row=r_idx, column=14).value = u["dlPhoto"]
            if u["punchCard"]:
                ws.cell(row=r_idx, column=15).value = u["punchCard"]
                
            updated = True
            break
            
    if updated:
        wb.save(EXCEL_FILE)
    wb.close()
    return updated, old_dl, old_pc

# Helper queries for Drivers
def get_drivers():
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Drivers"]
        drivers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: # Skip empty rows
                continue
            drivers.append({
                "punchNumber": str(row[0]).strip(),
                "driverName": str(row[1]).strip() if row[1] else "",
                "dlNumber": str(row[2]).strip() if row[2] else "",
                "panNumber": str(row[3]).strip() if row[3] else "",
                "accountNumber": str(row[4]).strip() if row[4] else "",
                "dlPhoto": str(row[5]).strip() if row[5] else "",
                "punchCard": str(row[6]).strip() if row[6] else "",
                "panPhoto": str(row[7]).strip() if row[7] else "",
                "passbookPhoto": str(row[8]).strip() if row[8] else "",
                "baseDocument": str(row[9]).strip() if row[9] else "",
                "status": str(row[10]).strip() if row[10] else "Inactive",
                "createdAt": str(row[11]).strip() if row[11] else ""
            })
        wb.close()
        return drivers
    except Exception as e:
        print(f"Error reading drivers Excel: {e}")
        return []

def add_driver(d):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Drivers"]
    ws.append([
        d["punchNumber"],
        d["driverName"],
        d["dlNumber"],
        d["panNumber"],
        d["accountNumber"],
        d["dlPhoto"],
        d["punchCard"],
        d["panPhoto"],
        d["passbookPhoto"],
        d["baseDocument"],
        d.get("status", "Inactive"),
        d["createdAt"]
    ])
    wb.save(EXCEL_FILE)
    wb.close()

def delete_driver(punch_no):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Drivers"]
    row_idx = -1
    dl_photo = ""
    punch_card = ""
    pan_photo = ""
    passbook_photo = ""
    base_doc = ""
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().upper() == punch_no.strip().upper():
            row_idx = r_idx
            dl_photo = ws.cell(row=r_idx, column=6).value or ""
            punch_card = ws.cell(row=r_idx, column=7).value or ""
            pan_photo = ws.cell(row=r_idx, column=8).value or ""
            passbook_photo = ws.cell(row=r_idx, column=9).value or ""
            base_doc = ws.cell(row=r_idx, column=10).value or ""
            break
            
    if row_idx != -1:
        ws.delete_rows(row_idx)
        wb.save(EXCEL_FILE)
        wb.close()
        return True, dl_photo, punch_card, pan_photo, passbook_photo, base_doc
    wb.close()
    return False, "", "", "", "", ""

def toggle_driver_status(punch_no):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Drivers"]
    updated = False
    new_val = ""
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().upper() == punch_no.strip().upper():
            current_val = ws.cell(row=r_idx, column=11).value or "Inactive"
            new_val = "Inactive" if current_val.strip() == "Active" else "Active"
            ws.cell(row=r_idx, column=11).value = new_val
            updated = True
            break
    if updated:
        wb.save(EXCEL_FILE)
    wb.close()
    return updated, new_val

# Helper queries for Users
def get_users():
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Users"]
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            users.append({
                "username": str(row[0]).strip(),
                "passwordHash": str(row[1]).strip() if row[1] else "",
                "role": str(row[2]).strip() if row[2] else "",
                "status": str(row[3]).strip() if row[3] else "",
                "createdAt": str(row[4]).strip() if row[4] else ""
            })
        wb.close()
        return users
    except Exception as e:
        print(f"Error reading users Excel: {e}")
        return []

def add_user(u):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    ws.append([
        u["username"],
        u["passwordHash"],
        u["role"],
        u["status"],
        u["createdAt"]
    ])
    wb.save(EXCEL_FILE)
    wb.close()

def update_user_status(username, status):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    updated = False
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().lower() == username.strip().lower():
            ws.cell(row=r_idx, column=4).value = status
            updated = True
            break
    if updated:
        wb.save(EXCEL_FILE)
    wb.close()
    return updated

def update_user_role(username, role):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    updated = False
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().lower() == username.strip().lower():
            ws.cell(row=r_idx, column=3).value = role
            updated = True
            break
    if updated:
        wb.save(EXCEL_FILE)
    wb.close()
    return updated

def delete_user_from_excel(username):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    row_idx = -1
    for r_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r_idx, column=1).value
        if cell_val and str(cell_val).strip().lower() == username.strip().lower():
            row_idx = r_idx
            break
    if row_idx != -1:
        ws.delete_rows(row_idx)
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    wb.close()
    return False

# Security Decorators
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
            
        # Verify user exists and is approved
        users = get_users()
        user = next((u for u in users if u['username'].lower() == session['username'].lower()), None)
        if not user or user['status'] != 'approved':
            session.clear()
            return jsonify({"error": "Account not approved or disabled."}), 403
            
        return f(*args, **kwargs)
    return decorated_function

def editor_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
            
        # Verify user has edit rights (admin or editor) and is approved
        users = get_users()
        user = next((u for u in users if u['username'].lower() == session['username'].lower()), None)
        if not user or user['status'] != 'approved' or user['role'] not in ['admin', 'editor']:
            return jsonify({"error": "Forbidden. Write/Edit permissions required."}), 403
            
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return jsonify({"error": "Unauthorized. Please log in."}), 401
            
        # Verify user is admin and approved
        users = get_users()
        user = next((u for u in users if u['username'].lower() == session['username'].lower()), None)
        if not user or user['status'] != 'approved' or user['role'] != 'admin':
            return jsonify({"error": "Forbidden. Administrator access required."}), 403
            
        return f(*args, **kwargs)
    return decorated_function

# --- API Routes ---

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

# -- Authentication Endpoints --

# 1. Login user
@app.route('/api/auth/login', methods=['POST'])
def api_login():
    try:
        data = request.json or {}
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({"error": "Username and password are required."}), 400
            
        users = get_users()
        user = next((u for u in users if u['username'].lower() == username.lower()), None)
        
        if not user or user['passwordHash'] != hash_password(password):
            return jsonify({"error": "Invalid username or password."}), 401
            
        if user['status'] == 'pending':
            return jsonify({"error": "Your account is pending admin approval.", "status": "pending"}), 403
        elif user['status'] == 'rejected':
            return jsonify({"error": "Your account approval has been rejected.", "status": "rejected"}), 403
            
        # Set session details
        session['username'] = user['username']
        session['role'] = user['role']
        session.permanent = True
        
        return jsonify({
            "success": True,
            "message": "Logged in successfully.",
            "user": {
                "username": user['username'],
                "role": user['role'],
                "status": user['status']
            }
        })
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({"error": "An error occurred during login."}), 500

# 2. Register user
@app.route('/api/auth/register', methods=['POST'])
def api_register():
    try:
        data = request.json or {}
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({"error": "Username and password are required."}), 400
            
        if len(username) < 3:
            return jsonify({"error": "Username must be at least 3 characters long."}), 400
        if len(password) < 4:
            return jsonify({"error": "Password must be at least 4 characters long."}), 400
            
        users = get_users()
        if any(u['username'].lower() == username.lower() for u in users):
            return jsonify({"error": f"Username '{username}' already exists."}), 400
            
        # Default registered user is Viewer with pending status
        role = "viewer"
        status = "pending"
        
        new_user = {
            "username": username,
            "passwordHash": hash_password(password),
            "role": role,
            "status": status,
            "createdAt": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        add_user(new_user)
        return jsonify({
            "success": True,
            "message": "Registration successful. Please wait for admin approval."
        })
    except Exception as e:
        print(f"Registration error: {e}")
        return jsonify({"error": "An error occurred during registration."}), 500

# 3. Get currently logged-in user profile
@app.route('/api/auth/me', methods=['GET'])
def api_me():
    if 'username' not in session:
        return jsonify({"authenticated": False}), 200
        
    users = get_users()
    user = next((u for u in users if u['username'].lower() == session['username'].lower()), None)
    
    if not user:
        session.clear()
        return jsonify({"authenticated": False}), 200
        
    return jsonify({
        "authenticated": True,
        "username": user['username'],
        "role": user['role'],
        "status": user['status']
    })

# 4. Logout user
@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({"success": True, "message": "Logged out successfully."})


# -- Vehicle Registry Endpoints (Access Guarded) --

# 1. Fetch all vehicles (All approved users)
@app.route('/api/vehicles', methods=['GET'])
@login_required
def api_get_vehicles():
    try:
        vehicles = get_vehicles()
        today = datetime.date.today()
        
        processed = []
        for v in vehicles:
            expired_docs = []
            expiring_soon_docs = []
            
            dates = {
                "Gatepass": v["gatepassExpiry"],
                "Insurance": v["insuranceExpiry"],
                "PUC": v["pucExpiry"],
                "Fitness": v["fitnessExpiry"],
                "Tax": v["taxExpiry"],
                "Permit": v["permitExpiry"]
            }
            
            for name, dt_str in dates.items():
                if dt_str:
                    try:
                        date_only_str = dt_str.split(' ')[0]
                        dt = datetime.datetime.strptime(date_only_str, "%Y-%m-%d").date()
                        if dt < today:
                            expired_docs.append({"name": name, "date": date_only_str})
                        elif (dt - today).days <= 15:
                            expiring_soon_docs.append({
                                "name": name,
                                "date": date_only_str,
                                "daysLeft": (dt - today).days
                            })
                    except ValueError as ve:
                        print(f"Date format mismatch for {name}: {dt_str}. Error: {ve}")
                        pass
            
            status = "Active"
            if expired_docs:
                status = "Expired"
            elif expiring_soon_docs:
                status = "Warning"
                
            v["status"] = status
            v["expiredDocs"] = expired_docs
            v["expiringSoonDocs"] = expiring_soon_docs
            processed.append(v)
            
        return jsonify(processed)
    except Exception as e:
        print(f"Error fetching vehicles: {e}")
        return jsonify({"error": "Failed to fetch vehicles."}), 500

# 2. Add vehicle details and file attachments (Admin & Editor Only)
@app.route('/api/vehicles', methods=['POST'])
@editor_required
def api_add_vehicle():
    try:
        vehicle_no = request.form.get('vehicleNo', '').strip().upper()
        if not vehicle_no:
            return jsonify({"error": "Vehicle Number is required."}), 400
            
        # Check if vehicle exists
        vehicles = get_vehicles()
        if any(v['vehicleNo'].upper() == vehicle_no for v in vehicles):
            return jsonify({"error": f"Vehicle {vehicle_no} is already registered."}), 400
            
        # Handle file uploads
        dl_photo = request.files.get('dlPhoto')
        punch_card = request.files.get('punchCard')
        
        dl_filename = ""
        pc_filename = ""
        
        # Clean vehicle number for safe filename
        safe_vehicle = "".join(c for c in vehicle_no if c.isalnum() or c == '-')
        timestamp = int(datetime.datetime.now().timestamp())
        
        if dl_photo and dl_photo.filename:
            ext = os.path.splitext(dl_photo.filename)[1]
            dl_filename = f"{safe_vehicle}_dl_{timestamp}{ext}"
            dl_photo.save(os.path.join(UPLOAD_FOLDER, dl_filename))

        if punch_card and punch_card.filename:
            ext = os.path.splitext(punch_card.filename)[1]
            pc_filename = f"{safe_vehicle}_pc_{timestamp}{ext}"
            punch_card.save(os.path.join(UPLOAD_FOLDER, pc_filename))
            
        # Add entry
        vehicle = {
            "vehicleNo": vehicle_no,
            "driverName": request.form.get('driverName', '').strip(),
            "punchNumber": request.form.get('punchNumber', '').strip(),
            "gatepassExpiry": request.form.get('gatepassExpiry', ''),
            "insuranceExpiry": request.form.get('insuranceExpiry', ''),
            "pucExpiry": request.form.get('pucExpiry', ''),
            "fitnessExpiry": request.form.get('fitnessExpiry', ''),
            "taxExpiry": request.form.get('taxExpiry', ''),
            "permitExpiry": request.form.get('permitExpiry', ''),
            "stageInspection": request.form.get('stageInspection', 'Pending').strip(),
            "stageSafety": request.form.get('stageSafety', 'Pending').strip(),
            "stageDept": request.form.get('stageDept', 'Pending').strip(),
            "remark": request.form.get('remark', '').strip(),
            "dlPhoto": dl_filename,
            "punchCard": pc_filename,
            "createdAt": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        add_vehicle(vehicle)
        return jsonify({"success": True, "message": "Vehicle added successfully."})
        
    except Exception as e:
        print(f"Error adding vehicle: {e}")
        return jsonify({"error": "Failed to add vehicle."}), 500

# 3. Update existing vehicle details (Admin & Editor Only)
@app.route('/api/vehicles/<vehicle_no>/update', methods=['POST'])
@editor_required
def api_update_vehicle(vehicle_no):
    try:
        vehicle_no = vehicle_no.strip().upper()
        
        # Check if vehicle exists
        vehicles = get_vehicles()
        existing = next((v for v in vehicles if v['vehicleNo'].upper() == vehicle_no), None)
        if not existing:
            return jsonify({"error": "Vehicle not found."}), 404
            
        # Handle file uploads
        dl_photo = request.files.get('dlPhoto')
        punch_card = request.files.get('punchCard')
        
        dl_filename = ""
        pc_filename = ""
        
        safe_vehicle = "".join(c for c in vehicle_no if c.isalnum() or c == '-')
        timestamp = int(datetime.datetime.now().timestamp())
        
        if dl_photo and dl_photo.filename:
            ext = os.path.splitext(dl_photo.filename)[1]
            dl_filename = f"{safe_vehicle}_dl_{timestamp}{ext}"
            dl_photo.save(os.path.join(UPLOAD_FOLDER, dl_filename))

        if punch_card and punch_card.filename:
            ext = os.path.splitext(punch_card.filename)[1]
            pc_filename = f"{safe_vehicle}_pc_{timestamp}{ext}"
            punch_card.save(os.path.join(UPLOAD_FOLDER, pc_filename))
            
        # Parse fields from form
        u = {
            "driverName": request.form.get('driverName', '').strip(),
            "punchNumber": request.form.get('punchNumber', '').strip(),
            "gatepassExpiry": request.form.get('gatepassExpiry', ''),
            "insuranceExpiry": request.form.get('insuranceExpiry', ''),
            "pucExpiry": request.form.get('pucExpiry', ''),
            "fitnessExpiry": request.form.get('fitnessExpiry', ''),
            "taxExpiry": request.form.get('taxExpiry', ''),
            "permitExpiry": request.form.get('permitExpiry', ''),
            "stageInspection": request.form.get('stageInspection', existing.get('stageInspection', 'Pending')).strip(),
            "stageSafety": request.form.get('stageSafety', existing.get('stageSafety', 'Pending')).strip(),
            "stageDept": request.form.get('stageDept', existing.get('stageDept', 'Pending')).strip(),
            "remark": request.form.get('remark', '').strip(),
            "dlPhoto": dl_filename,
            "punchCard": pc_filename
        }
        
        success, old_dl, old_pc = update_vehicle_in_excel(vehicle_no, u)
        
        if success:
            # Delete old attachments if replaced
            if dl_filename and old_dl:
                old_path = os.path.join(UPLOAD_FOLDER, old_dl)
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception as fe:
                        print(f"Failed to delete old DL photo: {fe}")
            if pc_filename and old_pc:
                old_path = os.path.join(UPLOAD_FOLDER, old_pc)
                if os.path.exists(old_path):
                    try:
                        os.remove(old_path)
                    except Exception as fe:
                        print(f"Failed to delete old punch card: {fe}")
                        
            return jsonify({"success": True, "message": "Vehicle details updated successfully."})
        else:
            return jsonify({"error": "Failed to update vehicle in spreadsheet."}), 500
            
    except Exception as e:
        print(f"Error updating vehicle: {e}")
        return jsonify({"error": "Failed to update vehicle details."}), 500

# 4. Delete vehicle and delete its files from the uploads directory (Admin & Editor Only)
@app.route('/api/vehicles/<vehicle_no>', methods=['DELETE'])
@editor_required
def api_delete_vehicle(vehicle_no):
    try:
        vehicle_no = vehicle_no.strip().upper()
        success, dl_photo, punch_card = delete_vehicle(vehicle_no)
        
        if success:
            # Delete attachments
            if dl_photo:
                dl_path = os.path.join(UPLOAD_FOLDER, dl_photo)
                if os.path.exists(dl_path):
                    try:
                        os.remove(dl_path)
                    except Exception as fe:
                        print(f"Failed to delete file {dl_photo}: {fe}")
            if punch_card:
                pc_path = os.path.join(UPLOAD_FOLDER, punch_card)
                if os.path.exists(pc_path):
                    try:
                        os.remove(pc_path)
                    except Exception as fe:
                        print(f"Failed to delete file {punch_card}: {fe}")
            return jsonify({"success": True, "message": f"Vehicle {vehicle_no} removed successfully."})
        else:
            return jsonify({"error": "Vehicle not found."}), 404
            
    except Exception as e:
        print(f"Error deleting vehicle: {e}")
        return jsonify({"error": "Failed to delete vehicle."}), 500

# 5. Download Excel Database (Admin & Editor Only)
@app.route('/api/download')
@editor_required
def api_download():
    try:
        init_excel()
        if os.path.exists(EXCEL_FILE):
            return send_file(EXCEL_FILE, as_attachment=True, download_name='gatepass_data.xlsx')
        return jsonify({"error": "Excel database file not initialized."}), 404
    except Exception as e:
        print(f"Error downloading file: {e}")
        return jsonify({"error": "Failed to download Excel file."}), 500

# 6. Toggle a specific gatepass clearance stage status between Pending and Done (Admin & Editor Only)
@app.route('/api/vehicles/<vehicle_no>/toggle-stage', methods=['POST'])
@editor_required
def api_toggle_vehicle_stage(vehicle_no):
    try:
        vehicle_no = vehicle_no.strip().upper()
        data = request.json or {}
        stage_name = data.get('stage', '').strip() # "Inspection", "Safety", "Department"
        
        if stage_name not in ["Inspection", "Safety", "Department"]:
            return jsonify({"error": "Invalid stage name. Must be Inspection, Safety, or Department."}), 400
            
        success, new_val = toggle_vehicle_stage(vehicle_no, stage_name)
        if success:
            return jsonify({
                "success": True, 
                "message": f"Vehicle '{vehicle_no}' stage '{stage_name}' toggled to '{new_val}'.",
                "newValue": new_val
            })
        return jsonify({"error": "Vehicle not found."}), 404
    except Exception as e:
        print(f"Error toggling vehicle stage: {e}")
        return jsonify({"error": "Failed to toggle stage status."}), 500


# -- Driver & Punch Card Endpoints (Access Guarded) --

# 1. Fetch all driver profiles
@app.route('/api/drivers', methods=['GET'])
@login_required
def api_get_drivers():
    try:
        drivers = get_drivers()
        return jsonify(drivers)
    except Exception as e:
        print(f"Error fetching drivers: {e}")
        return jsonify({"error": "Failed to fetch drivers."}), 500

# 2. Add driver profile and documents (DL, Punch Card, PAN, Passbook, Base Doc)
@app.route('/api/drivers', methods=['POST'])
@editor_required
def api_add_driver():
    try:
        punch_no = request.form.get('punchNumber', '').strip().upper()
        driver_name = request.form.get('driverName', '').strip()
        
        if not punch_no or not driver_name:
            return jsonify({"error": "Driver Name and Punch Number are required."}), 400
            
        # Check if punch card already exists in the system
        drivers = get_drivers()
        if any(d['punchNumber'].upper() == punch_no for d in drivers):
            return jsonify({"error": f"Punch Card {punch_no} is already registered."}), 400
            
        # Handle file uploads
        dl_photo = request.files.get('dlPhoto')
        punch_card = request.files.get('punchCard')
        pan_photo = request.files.get('panPhoto')
        passbook_photo = request.files.get('passbookPhoto')
        base_doc = request.files.get('baseDocument')
        
        # DL and Punch Card photo uploads are strictly required
        if not dl_photo or not dl_photo.filename:
            return jsonify({"error": "Driver's License (DL) photo upload is required."}), 400
        if not punch_card or not punch_card.filename:
            return jsonify({"error": "Punch Card photo upload is required."}), 400
            
        dl_filename = ""
        pc_filename = ""
        pan_filename = ""
        pb_filename = ""
        bd_filename = ""
        
        safe_punch = "".join(c for c in punch_no if c.isalnum() or c == '-')
        timestamp = int(datetime.datetime.now().timestamp())
        
        if dl_photo and dl_photo.filename:
            ext = os.path.splitext(dl_photo.filename)[1]
            dl_filename = f"driver_{safe_punch}_dl_{timestamp}{ext}"
            dl_photo.save(os.path.join(UPLOAD_FOLDER, dl_filename))

        if punch_card and punch_card.filename:
            ext = os.path.splitext(punch_card.filename)[1]
            pc_filename = f"driver_{safe_punch}_pc_{timestamp}{ext}"
            punch_card.save(os.path.join(UPLOAD_FOLDER, pc_filename))

        if pan_photo and pan_photo.filename:
            ext = os.path.splitext(pan_photo.filename)[1]
            pan_filename = f"driver_{safe_punch}_pan_{timestamp}{ext}"
            pan_photo.save(os.path.join(UPLOAD_FOLDER, pan_filename))

        if passbook_photo and passbook_photo.filename:
            ext = os.path.splitext(passbook_photo.filename)[1]
            pb_filename = f"driver_{safe_punch}_pb_{timestamp}{ext}"
            passbook_photo.save(os.path.join(UPLOAD_FOLDER, pb_filename))

        if base_doc and base_doc.filename:
            ext = os.path.splitext(base_doc.filename)[1]
            bd_filename = f"driver_{safe_punch}_basedoc_{timestamp}{ext}"
            base_doc.save(os.path.join(UPLOAD_FOLDER, bd_filename))
            
        driver = {
            "punchNumber": punch_no,
            "driverName": driver_name,
            "dlNumber": request.form.get('dlNumber', '').strip(),
            "panNumber": request.form.get('panNumber', '').strip(),
            "accountNumber": request.form.get('accountNumber', '').strip(),
            "dlPhoto": dl_filename,
            "punchCard": pc_filename,
            "panPhoto": pan_filename,
            "passbookPhoto": pb_filename,
            "baseDocument": bd_filename,
            "status": request.form.get('status', 'Inactive').strip(),
            "createdAt": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        add_driver(driver)
        return jsonify({"success": True, "message": "Driver profile saved successfully."})
    except Exception as e:
        print(f"Error adding driver: {e}")
        return jsonify({"error": "Failed to save driver profile."}), 500

# 3. Toggle driver status (Active/Inactive)
@app.route('/api/drivers/<punch_no>/toggle-status', methods=['POST'])
@editor_required
def api_toggle_driver_status(punch_no):
    try:
        punch_no = punch_no.strip().upper()
        success, new_val = toggle_driver_status(punch_no)
        if success:
            return jsonify({
                "success": True,
                "message": f"Punch card status updated to {new_val}.",
                "newValue": new_val
            })
        return jsonify({"error": "Driver profile not found."}), 404
    except Exception as e:
        print(f"Error toggling driver status: {e}")
        return jsonify({"error": "Failed to update driver status."}), 500

# 4. Delete driver profile and documents
@app.route('/api/drivers/<punch_no>', methods=['DELETE'])
@editor_required
def api_delete_driver(punch_no):
    try:
        punch_no = punch_no.strip().upper()
        success, dl_photo, punch_card, pan_photo, passbook_photo, base_doc = delete_driver(punch_no)
        
        if success:
            # Delete attachments
            for filename in [dl_photo, punch_card, pan_photo, passbook_photo, base_doc]:
                if filename:
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    if os.path.exists(filepath):
                        try:
                            os.remove(filepath)
                        except Exception as fe:
                            print(f"Failed to delete file {filename}: {fe}")
            return jsonify({"success": True, "message": f"Driver profile {punch_no} removed successfully."})
        else:
            return jsonify({"error": "Driver profile not found."}), 404
    except Exception as e:
        print(f"Error deleting driver profile: {e}")
        return jsonify({"error": "Failed to remove driver profile."}), 500


# -- User Approval & Access Management (Admin Only) --

# 1. Get all registered users
@app.route('/api/admin/users', methods=['GET'])
@admin_required
def api_admin_users():
    try:
        users = get_users()
        safe_users = []
        for u in users:
            safe_users.append({
                "username": u["username"],
                "role": u["role"],
                "status": u["status"],
                "createdAt": u["createdAt"]
            })
        return jsonify(safe_users)
    except Exception as e:
        print(f"Admin fetch users error: {e}")
        return jsonify({"error": "Failed to retrieve user registry."}), 500

# 2. Update user approval status (Approve/Reject)
@app.route('/api/admin/users/<username>/status', methods=['POST'])
@admin_required
def api_admin_update_status(username):
    try:
        username = username.strip()
        data = request.json or {}
        status = data.get('status', '').strip().lower()
        
        if status not in ['approved', 'rejected', 'pending']:
            return jsonify({"error": "Invalid status value."}), 400
            
        if username.lower() == 'admin':
            return jsonify({"error": "Cannot modify main admin account status."}), 400
            
        success = update_user_status(username, status)
        if success:
            return jsonify({"success": True, "message": f"User '{username}' status updated to '{status}'."})
        return jsonify({"error": "User not found."}), 404
    except Exception as e:
        print(f"Admin update status error: {e}")
        return jsonify({"error": "Failed to update user status."}), 500

# 3. Update user access role (Admin/Editor/Viewer)
@app.route('/api/admin/users/<username>/role', methods=['POST'])
@admin_required
def api_admin_update_role(username):
    try:
        username = username.strip()
        data = request.json or {}
        role = data.get('role', '').strip().lower()
        
        if role not in ['admin', 'editor', 'viewer']:
            return jsonify({"error": "Invalid role value."}), 400
            
        if username.lower() == 'admin':
            return jsonify({"error": "Cannot modify main admin account role."}), 400
            
        success = update_user_role(username, role)
        if success:
            return jsonify({"success": True, "message": f"User '{username}' role updated to '{role}'."})
        return jsonify({"error": "User not found."}), 404
    except Exception as e:
        print(f"Admin update role error: {e}")
        return jsonify({"error": "Failed to update user role."}), 500

# 4. Delete user account
@app.route('/api/admin/users/<username>', methods=['DELETE'])
@admin_required
def api_admin_delete_user(username):
    try:
        username = username.strip()
        if username.lower() == 'admin':
            return jsonify({"error": "Cannot delete main admin account."}), 400
            
        success = delete_user_from_excel(username)
        if success:
            return jsonify({"success": True, "message": f"User '{username}' deleted successfully."})
        return jsonify({"error": "User not found."}), 404
    except Exception as e:
        print(f"Admin delete user error: {e}")
        return jsonify({"error": "Failed to delete user."}), 500


if __name__ == '__main__':
    init_excel()
    # Run server on port 5000
    app.run(host='0.0.0.0', port=5000, debug=True)
